import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.backends.backend_pdf import PdfPages
except ImportError:
    plt = None

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

DATA_FILE = Path(__file__).with_name("vendas.csv")


def carregar_dados(path):
    registros = []
    if not path.exists():
        raise FileNotFoundError(f"Arquivo de dados não encontrado: {path}")

    with open(path, encoding="utf-8", newline="") as arquivo:
        leitor = csv.DictReader(arquivo)
        for linha in leitor:
            try:
                quantidade = int(linha["quantidade"])
                preco_unitario = float(linha["preco_unitario"])
            except (ValueError, TypeError):
                continue

            data_text = linha.get("data") or linha.get("data_venda")
            data_obj = None
            try:
                data_obj = datetime.strptime(data_text, "%Y-%m-%d").date()
            except (TypeError, ValueError):
                pass

            registros.append({
                "data_venda": data_text,
                "data": data_obj,
                "produto": linha.get("produto"),
                "categoria": linha.get("categoria"),
                "quantidade": quantidade,
                "preco_unitario": preco_unitario,
                "regiao": linha.get("regiao"),
                "receita": quantidade * preco_unitario,
            })
    return registros


def calcular_métricas(registros):
    total_registros = len(registros)
    receita_total = sum(registro["receita"] for registro in registros)
    vendas_eletronicos = [r for r in registros if r["categoria"] == "Eletrônicos"]

    produtos = Counter()
    regiao_valores = defaultdict(float)
    receita_categoria = defaultdict(float)
    receita_mensal = defaultdict(float)
    pivot_regiao_categoria = defaultdict(lambda: defaultdict(float))

    for registro in registros:
        produtos[registro["produto"]] += registro["quantidade"]
        regiao_valores[registro["regiao"]] += registro["receita"]
        receita_categoria[registro["categoria"]] += registro["receita"]
        pivot_regiao_categoria[registro["regiao"]][registro["categoria"]] += registro["receita"]
        periodo = registro["data"].strftime("%Y-%m") if registro["data"] else registro["data_venda"][:7]
        receita_mensal[periodo] += registro["receita"]

    produto_top, quantidade_top = produtos.most_common(1)[0] if produtos else (None, 0)
    regiao_top = max(regiao_valores, key=regiao_valores.get) if regiao_valores else None
    valor_regiao_top = regiao_valores.get(regiao_top, 0.0)

    return {
        "total_registros": total_registros,
        "receita_total": receita_total,
        "vendas_eletronicos": vendas_eletronicos,
        "produto_top": produto_top,
        "quantidade_top": quantidade_top,
        "regiao_top": regiao_top,
        "valor_regiao_top": valor_regiao_top,
        "receita_categoria": dict(sorted(receita_categoria.items(), key=lambda item: item[1], reverse=True)),
        "top_produtos_quantidade": dict(produtos.most_common(10)),
        "receita_regiao": dict(sorted(regiao_valores.items(), key=lambda item: item[1], reverse=True)),
        "receita_mensal": dict(sorted(receita_mensal.items())),
        "pivot_regiao_categoria": {
            regiao: dict(categorias)
            for regiao, categorias in pivot_regiao_categoria.items()
        },
    }


def criar_treeview(frame, colunas, largura_coluna, dados):
    tree = ttk.Treeview(frame, columns=colunas, show="headings", height=6)
    for coluna, largura in zip(colunas, largura_coluna):
        tree.heading(coluna, text=coluna)
        tree.column(coluna, width=largura, anchor=tk.W)
    for item in dados:
        tree.insert("", tk.END, values=item)
    scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    return tree


def criar_grafico(frame, dados, titulo, xlabel, ylabel, rotacao=0, tipo="bar"):
    if plt is None:
        label = ttk.Label(frame, text="Matplotlib não está disponível.")
        label.pack(fill=tk.BOTH, expand=True)
        return

    figura, ax = plt.subplots(figsize=(5, 4), dpi=100)
    chaves = list(dados.keys())
    valores = list(dados.values())
    if tipo == "line":
        ax.plot(chaves, valores, marker="o", color="#4c72b0")
    else:
        ax.bar(chaves, valores, color="#4c72b0")
    ax.set_title(titulo)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.tick_params(axis="x", rotation=rotacao)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    figura.tight_layout()

    canvas = FigureCanvasTkAgg(figura, master=frame)
    canvas.draw()
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack(fill=tk.BOTH, expand=True)


def montar_pivot_rows(pivot):
    categorias = sorted(
        {categoria for categorias in pivot.values() for categoria in categorias}
    )
    colunas = ["Região"] + categorias
    rows = []
    for regiao in sorted(pivot.keys(), key=lambda reg: sum(pivot[reg].values()), reverse=True):
        linha = [regiao] + [f"R$ {pivot[regiao].get(categoria, 0):,.2f}" for categoria in categorias]
        rows.append(tuple(linha))
    return colunas, rows


def exportar_para_xlsx(metricas, registros, caminho):
    if Workbook is None:
        raise ImportError("openpyxl não está instalado.")

    workbook = Workbook()
    resumo = workbook.active
    resumo.title = "Resumo"
    linhas = [
        ("Métrica", "Valor"),
        ("Total de registros", metricas["total_registros"]),
        ("Receita total", f"R$ {metricas['receita_total']:,.2f}"),
        ("Produto mais vendido", f"{metricas['produto_top']} ({metricas['quantidade_top']} unidades)"),
        ("Região com maior valor", f"{metricas['regiao_top']} (R$ {metricas['valor_regiao_top']:,.2f})"),
    ]
    for linha in linhas:
        resumo.append(linha)

    pivot_sheet = workbook.create_sheet(title="Pivot Região x Categoria")
    colunas_pivot, linhas_pivot = montar_pivot_rows(metricas["pivot_regiao_categoria"])
    pivot_sheet.append(colunas_pivot)
    for linha in linhas_pivot:
        pivot_sheet.append(linha)

    categoria_sheet = workbook.create_sheet(title="Receita por Categoria")
    categoria_sheet.append(("Categoria", "Receita"))
    for categoria, valor in metricas["receita_categoria"].items():
        categoria_sheet.append((categoria, valor))

    mensal_sheet = workbook.create_sheet(title="Receita Mensal")
    mensal_sheet.append(("Período", "Receita"))
    for periodo, valor in metricas["receita_mensal"].items():
        mensal_sheet.append((periodo, valor))

    largura_colunas = {}
    for planilha in workbook.worksheets:
        for linha in planilha.iter_rows(values_only=True):
            for idx, valor in enumerate(linha, start=1):
                largura_colunas[idx] = max(largura_colunas.get(idx, 0), len(str(valor)))
        for idx, largura in largura_colunas.items():
            planilha.column_dimensions[get_column_letter(idx)].width = largura + 2
        largura_colunas.clear()

    workbook.save(caminho)


def exportar_para_pdf(metricas, caminho):
    if plt is None:
        raise ImportError("matplotlib não está disponível para gerar PDF.")

    with PdfPages(caminho) as pdf:
        figura = plt.figure(figsize=(8.27, 11.69), dpi=100)
        figura.clf()
        texto = (
            f"Resumo de Vendas\n\n"
            f"Total de registros: {metricas['total_registros']}\n"
            f"Receita total: R$ {metricas['receita_total']:,.2f}\n"
            f"Produto mais vendido: {metricas['produto_top']} ({metricas['quantidade_top']} unidades)\n"
            f"Região com maior valor: {metricas['regiao_top']} (R$ {metricas['valor_regiao_top']:,.2f})\n"
        )
        figura.text(0.1, 0.95, texto, fontsize=12, va="top")
        pdf.savefig(figura)

        def salvar_figura(dados, titulo, xlabel, ylabel, tipo="bar"):
            fig, ax = plt.subplots(figsize=(8, 4.5), dpi=100)
            chaves = list(dados.keys())
            valores = list(dados.values())
            if tipo == "line":
                ax.plot(chaves, valores, marker="o", color="#4c72b0")
            else:
                ax.bar(chaves, valores, color="#4c72b0")
            ax.set_title(titulo)
            ax.set_xlabel(xlabel)
            ax.set_ylabel(ylabel)
            ax.tick_params(axis="x", rotation=45)
            ax.grid(axis="y", linestyle="--", alpha=0.4)
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)

        salvar_figura(metricas["receita_categoria"], "Receita por categoria", "Categoria", "Receita (R$)")
        salvar_figura(metricas["receita_mensal"], "Evolução mensal de receita", "Período", "Receita (R$)", tipo="line")

        pivot = metricas["pivot_regiao_categoria"]
        categorias = sorted({cat for categorias in pivot.values() for cat in categorias})
        linhas = [tuple(["Região"] + categorias)]
        for regiao in sorted(pivot.keys(), key=lambda reg: sum(pivot[reg].values()), reverse=True):
            linhas.append(tuple([regiao] + [pivot[regiao].get(categoria, 0) for categoria in categorias]))
        fig, ax = plt.subplots(figsize=(8.27, 6), dpi=100)
        ax.axis("tight")
        ax.axis("off")
        tabela = ax.table(cellText=linhas, loc="center", cellLoc="center")
        tabela.auto_set_font_size(False)
        tabela.set_fontsize(8)
        tabela.scale(1, 1.5)
        pdf.savefig(fig)
        plt.close(fig)


def salvar_relatorio_xlsx(root, metricas, registros):
    if Workbook is None:
        messagebox.showwarning(
            "Dependência ausente",
            "Instale openpyxl para exportar para XLSX: python -m pip install openpyxl",
        )
        return
    caminho = filedialog.asksaveasfilename(
        parent=root,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
    )
    if not caminho:
        return
    try:
        exportar_para_xlsx(metricas, registros, caminho)
        messagebox.showinfo("Exportação concluída", f"Relatório XLSX salvo em:\n{caminho}")
    except Exception as exc:
        messagebox.showerror("Erro ao exportar", str(exc))


def salvar_relatorio_pdf(root, metricas):
    if plt is None:
        messagebox.showwarning(
            "Dependência ausente",
            "Instale matplotlib para exportar para PDF.",
        )
        return
    caminho = filedialog.asksaveasfilename(
        parent=root,
        defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")],
    )
    if not caminho:
        return
    try:
        exportar_para_pdf(metricas, caminho)
        messagebox.showinfo("Exportação concluída", f"Relatório PDF salvo em:\n{caminho}")
    except Exception as exc:
        messagebox.showerror("Erro ao exportar", str(exc))


def criar_interface(registros, métricas):
    root = tk.Tk()
    root.title("Análise de Vendas - TechStore")
    root.geometry("1200x840")

    notebook = ttk.Notebook(root)
    notebook.pack(fill=tk.BOTH, expand=True)

    resumo_frame = ttk.Frame(notebook, padding=12)
    dados_frame = ttk.Frame(notebook, padding=12)
    graficos_frame = ttk.Frame(notebook, padding=12)
    relatorio_frame = ttk.Frame(notebook, padding=12)

    notebook.add(resumo_frame, text="Resumo")
    notebook.add(dados_frame, text="Dados")
    notebook.add(graficos_frame, text="Gráficos")
    notebook.add(relatorio_frame, text="Relatório")

    # Resumo
    metricas_frame = ttk.Frame(resumo_frame)
    metricas_frame.pack(fill=tk.X, pady=4)

    labels = [
        ("Total de registros:", métricas["total_registros"]),
        ("Receita total:", f"R$ {métricas['receita_total']:,.2f}"),
        ("Produto mais vendido:", f"{métricas['produto_top']} ({métricas['quantidade_top']} unidades)"),
        ("Região com maior valor:", f"{métricas['regiao_top']} (R$ {métricas['valor_regiao_top']:,.2f})"),
    ]

    for idx, (texto, valor) in enumerate(labels):
        label = ttk.Label(metricas_frame, text=f"{texto} {valor}", font=(None, 12, "bold"))
        label.grid(row=idx, column=0, sticky="w", pady=2)

    eletronicos_frame = ttk.LabelFrame(resumo_frame, text="Vendas Eletrônicos")
    eletronicos_frame.pack(fill=tk.BOTH, expand=True, pady=8)

    colunas_elec = ["data_venda", "produto", "quantidade", "preco_unitario", "receita", "regiao"]
    dados_eletronicos = [
        (
            registro["data_venda"],
            registro["produto"],
            registro["quantidade"],
            f"R$ {registro['preco_unitario']:,.2f}",
            f"R$ {registro['receita']:,.2f}",
            registro["regiao"],
        )
        for registro in métricas["vendas_eletronicos"]
    ]
    criar_treeview(eletronicos_frame, colunas_elec, [120, 180, 90, 120, 120, 120], dados_eletronicos)

    # Dados
    primeiros_frame = ttk.LabelFrame(dados_frame, text="Primeiras 5 linhas")
    primeiros_frame.pack(fill=tk.BOTH, expand=False, pady=8)

    colunas = ["data_venda", "produto", "categoria", "quantidade", "preco_unitario", "receita", "regiao"]
    dados_principais = [
        (
            registro["data_venda"],
            registro["produto"],
            registro["categoria"],
            registro["quantidade"],
            f"R$ {registro['preco_unitario']:,.2f}",
            f"R$ {registro['receita']:,.2f}",
            registro["regiao"],
        )
        for registro in registros[:5]
    ]
    criar_treeview(primeiros_frame, colunas, [100, 170, 120, 90, 120, 120, 120], dados_principais)

    # Relatório
    export_frame = ttk.Frame(relatorio_frame)
    export_frame.pack(fill=tk.X, pady=4)

    xlsx_button = ttk.Button(
        export_frame,
        text="Exportar para XLSX",
        command=lambda: salvar_relatorio_xlsx(root, métricas, registros),
    )
    pdf_button = ttk.Button(
        export_frame,
        text="Exportar para PDF",
        command=lambda: salvar_relatorio_pdf(root, métricas),
    )
    xlsx_button.pack(side=tk.LEFT, padx=4)
    pdf_button.pack(side=tk.LEFT, padx=4)

    pivot_frame = ttk.LabelFrame(relatorio_frame, text="Pivot Região × Categoria")
    pivot_frame.pack(fill=tk.BOTH, expand=True, pady=8)
    colunas_pivot, linhas_pivot = montar_pivot_rows(métricas["pivot_regiao_categoria"])
    criar_treeview(pivot_frame, colunas_pivot, [140] + [130] * (len(colunas_pivot) - 1), linhas_pivot)

    # Gráficos
    graficos_grid = ttk.Frame(graficos_frame)
    graficos_grid.pack(fill=tk.BOTH, expand=True)
    graficos_grid.columnconfigure(0, weight=1)
    graficos_grid.columnconfigure(1, weight=1)
    graficos_grid.rowconfigure(0, weight=1)
    graficos_grid.rowconfigure(1, weight=1)

    chart1_frame = ttk.LabelFrame(graficos_grid, text="Receita por categoria")
    chart2_frame = ttk.LabelFrame(graficos_grid, text="Top 10 produtos por quantidade")
    chart3_frame = ttk.LabelFrame(graficos_grid, text="Receita por região")
    chart4_frame = ttk.LabelFrame(graficos_grid, text="Receita mensal")

    chart1_frame.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
    chart2_frame.grid(row=0, column=1, sticky="nsew", padx=6, pady=6)
    chart3_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=6)
    chart4_frame.grid(row=1, column=1, sticky="nsew", padx=6, pady=6)

    criar_grafico(chart1_frame, métricas["receita_categoria"], "Receita por categoria", "Categoria", "Receita (R$)", rotacao=30)
    criar_grafico(chart2_frame, métricas["top_produtos_quantidade"], "Top 10 produtos mais vendidos", "Produto", "Quantidade", rotacao=45)
    criar_grafico(chart3_frame, métricas["receita_regiao"], "Receita por região", "Região", "Receita (R$)")
    criar_grafico(chart4_frame, métricas["receita_mensal"], "Evolução mensal de receita", "Período", "Receita (R$)", rotacao=45, tipo="line")

    root.mainloop()


def main():
    try:
        registros = carregar_dados(DATA_FILE)
    except FileNotFoundError as exc:
        messagebox.showerror("Erro", str(exc))
        return

    métrica = calcular_métricas(registros)
    criar_interface(registros, métrica)


if __name__ == "__main__":
    main()
